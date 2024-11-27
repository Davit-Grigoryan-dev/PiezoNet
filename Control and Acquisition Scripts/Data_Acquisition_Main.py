import pyvisa
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import GW_RIGOL
import time
import serial
import RIGOL_SCOPE
import xmlrpc.client as xml


plecs = xml.Server('http://localhost:1080/RPC2')

#GIT
# # Initialize the VISA resource manager
rm = pyvisa.ResourceManager() 
#rm1 = pyvisa.ResourceManager()  # Use the PyVISA-py backend
#timeout
rm.timeout = 20000



# Identify and set the resource name for your oscilloscope
ip_address = "192.168.0.100"
oscilloscope = RIGOL_SCOPE.init_oscilloscope(ip_address=ip_address)





def save_waveform_to_xlsx(waveform_ch1, waveform_ch2, sample_number, filename_prefix='ITEM', sig_figs=10):
    filename = f"{filename_prefix}_{sample_number}_waveform_data_T5-100C.xlsx"
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Waveform Data"
    
    # Find the next available columns
    max_col = sheet.max_column
    if max_col % 2 == 0:  # Ensure columns are grouped correctly
        next_col_ch1_voltage = max_col + 1
        next_col_ch2_voltage = max_col + 2
    else:
        next_col_ch1_voltage = max_col + (2 - max_col % 2)
        next_col_ch2_voltage = next_col_ch1_voltage + 1
    
    # Write data
    for row_num, (v1, v2) in enumerate(zip(waveform_ch1, waveform_ch2), start=1):
        sheet.cell(row=row_num, column=next_col_ch1_voltage, value=round(v1, sig_figs))
        sheet.cell(row=row_num, column=next_col_ch2_voltage, value=round(v2, sig_figs))
    
    workbook.save(filename)
    print(f"Waveform data saved to {filename}")



def calculate_efficiency(psu_voltage, psu_current, load_voltage, load_current):
    try:
        input_power = psu_voltage * psu_current
        output_power = load_voltage * load_current
        efficiency = (output_power / input_power) * 100 if input_power > 0 else 0
        return efficiency
    except ZeroDivisionError:
        return 0
    

def sweep_and_find_best_efficiency(start, end, step, psu, load, plecs):

    values = []
    highest_efficiency = 0
    best_value = None

    current_value = start
    while current_value <= end:
        # Set the current value to the system being tested
        plecs.plecs.set('control_system_6_submodules_60hz_Strobe/PHASE', 'Value', str(current_value))
        time.sleep(2)

        psu_voltage_readings = []
        psu_current_readings = []
        load_voltage_readings = []
        load_current_readings = []


              # Measure 10 times and take average
        for i in range(5):  # Average over 5 measurements
            load.write('*CLS')
            #oscilloscope.write('*CLS')
            psu_current_readings.append(float(psu.query('MEAS:CURR?')))
            time.sleep(0.2)
            load_current_readings.append(float(load.query(":MEAS:CURR:DC?")))
            time.sleep(0.2)
            psu_voltage_readings.append(float(psu.query('MEAS:VOLT?')))
            time.sleep(0.2)
            load_voltage_readings.append(float(load.query(":MEAS:VOLT:DC?")))
            time.sleep(0.2)

        # Calculate averages
        avg_psu_voltage = sum(psu_voltage_readings) / len(psu_voltage_readings) 
        avg_psu_current = sum(psu_current_readings) / len(psu_current_readings)
        avg_load_voltage = sum(load_voltage_readings) / len(load_voltage_readings)
        avg_load_current = sum(load_current_readings) / len(load_current_readings)


        # Calculate efficiency
        efficiency = calculate_efficiency(avg_psu_voltage, avg_psu_current, avg_load_voltage, avg_load_current)
        print(f"Current value: {current_value}, Efficiency: {efficiency}%")

        # Store the value and check if it's the highest efficiency found
        values.append((current_value, efficiency))
        if efficiency > highest_efficiency:
            highest_efficiency = efficiency
            best_value = current_value

        # Move to the next step
        current_value += step

    print(f"Highest Efficiency: {highest_efficiency}% at value: {best_value}")
    return best_value




def sweep_voltage(psu, load, start_voltage, stop_voltage, psu_step, load_step, delay, current_limit, mode, sample_number):
    try:
        GW_RIGOL.control_output(psu, "ON")
        GW_RIGOL.control_load_input(load, "ON")
        load.timeout = 20000
        load.read_termination = '\n'
        voltage = start_voltage
        psu_step_counter = 0

        while voltage <= stop_voltage:
            GW_RIGOL.setup_power_supply(psu, voltage, current_limit)
            print(f"Sweeping Power Supply Voltage: {voltage} V")

            load_voltage = 20

            if voltage < 115:
                threshold = 5
            else:
                threshold = 20

            while load_voltage <= (voltage / 2) - threshold:
                GW_RIGOL.setup_load(load, load_voltage, mode)
                print(f"Setting Load {mode.lower()} to: {load_voltage} (within {voltage} V)")

                plecs.plecs.set('control_system_6_submodules_60hz_Strobe/Vin', 'Value', str(voltage))
                time.sleep(0.5)
                FINAL_PHASE = sweep_and_find_best_efficiency(0.37, 0.52, 0.02, psu, load, plecs)
                time.sleep(0.5)
                plecs.plecs.set('control_system_6_submodules_60hz_Strobe/PHASE', 'Value', str(FINAL_PHASE))
                time.sleep(1)
                RIGOL_SCOPE.calculate_vertical_div(voltage, vertical_divisions=4, scale_factor=0.8, channel=3, instrument=oscilloscope)
                time.sleep(5)
                print("Getting waveform data...")
                time1, voltage_ch1 = RIGOL_SCOPE.get_waveform(oscilloscope, "CHAN3")
                time.sleep(1)
                time2, voltage_ch2 = RIGOL_SCOPE.get_waveform(oscilloscope, "CHAN2")
                oscilloscope.write(":RUN")
                PR_Voltage = voltage_ch1
                PR_Current = voltage_ch2 / 1.004  # Shunt resistor value

                save_waveform_to_xlsx(PR_Voltage, PR_Current, sample_number=sample_number, sig_figs=10)

                time.sleep(delay)
                load_voltage += load_step

            voltage += psu_step
            psu_step_counter += 1  # Increment the PSU step counter

        GW_RIGOL.control_output(psu, "OFF")
        GW_RIGOL.control_load_input(load, "OFF")
        RIGOL_SCOPE.close_oscilloscope(oscilloscope)

    except pyvisa.errors.VisaIOError as e:
        print(f"Error during voltage sweep: {e}")

# MAIN SECTION
#_____________________________________________________________________________________________________________

psu = None
load = None

if rm:
    com_port_psu = '3'
    usb_resource_load = 'USB0::0x1AB1::0x0E11::DL3A210700113::INSTR'  # Updated to the USB resource for the load
    
    psu = GW_RIGOL.initialize_power_supply(rm, com_port_psu)
    load = GW_RIGOL.initialize_load(rm, usb_resource_load)  # Use the USB resource for the load

    if psu and load:
        start_voltage = float(input("Enter start voltage: ").strip())
        stop_voltage = float(input("Enter stop voltage: ").strip())
        psu_step = float(input("Enter power supply step size: ").strip())
        load_step = float(input("Enter load step size: ").strip())
        delay = float(input("Enter delay between load steps (in seconds): ").strip())
        current_limit = float(input("Enter current limit: ").strip())
        mode = input("Select load mode (VOLTAGE, CURRENT, POWER): ").strip().upper()
        sample_number = input("Enter sample number: ").strip()

        if mode not in ["VOLTAGE", "CURRENT", "POWER"]:
            print("Invalid mode selected. Please select VOLTAGE, CURRENT, or POWER.")
        else:
            sweep_voltage(psu, load, start_voltage, stop_voltage, psu_step, load_step, delay, current_limit, mode, sample_number)